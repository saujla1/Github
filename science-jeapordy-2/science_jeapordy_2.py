import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook

page = requests.get("https://jeopardylabs.com/play/5th-grade-science-review80")
soup = BeautifulSoup(page.content, 'html.parser')
grid=soup.find('div', id='question-grid')
table=grid.find_all('div', class_='table-cell-inner')


#first_row=table[1]
print(type(table))
print((len(table)))
print(table)
question=[]
answer=[]
for rows in table:
    q=rows.find('div', class_='question').text
    question.append(q)
    
    a=rows.find('div', class_='answer').text
    answer.append(a)
    
#print(question)
#print(answer)
#row_1=soup.find_all('div', class_="table-cell-inner")
#print(row_1)
#q=grid.find_all('div', class_='table-row')
#print(q.text)
#a=q.find_all('div', class_='answer')
#print(a.text)

test_df = pd.DataFrame({'question': question,
                       'answer': answer,})
#print(test_df)
book = load_workbook('D:\\Projects\\Python\\science-jeapordy-2\\temp.xlsx')
sheet=book.worksheets[0]
row_count = sheet.max_row
print(row_count)
writer = pd.ExcelWriter('D:\\Projects\\Python\\science-jeapordy-2\\temp.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#test_df.to_excel(writer, "Main", startrow=len(test_df)+2)

test_df.to_excel(writer,header=False,startrow =row_count+2)
#test_df.to_excel(writer, index=False)
writer.save()
#test_df.to_excel('D:\\Projects\\Python\\science-jeapordy-2\\temp.xls')